"""Manipulação de planilhas Excel: reconstrução de linhas e aplicação de preços."""

import re
import io
import base64
from copy import copy
import openpyxl
from fastapi.responses import JSONResponse

from api.helpers import extract_dim_from_header, set_price_cell


def _get_code_str(item):
    """Extrai o código em string de um item que pode ser string ou objeto {code, ...}."""
    return item["code"] if isinstance(item, dict) else item


def rebuild_product_rows(ws, mapping):
    """Reorganiza as linhas de código de produto conforme o mapeamento do usuário."""
    formato_rows = []
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(row=r, column=1).value or "").strip().upper()
        dim = extract_dim_from_header(val)
        if dim is not None:
            formato_rows.append((r, val))

    for i in reversed(range(len(formato_rows))):
        hrow, header = formato_rows[i]
        if header not in mapping:
            continue

        user_codes = mapping[header].get("codes", [])
        if not user_codes:
            continue

        next_hrow = formato_rows[i + 1][0] if i + 1 < len(formato_rows) else ws.max_row + 1
        existing_code_rows = []
        for r in range(hrow + 1, next_hrow):
            cv = str(ws.cell(row=r, column=1).value or "").strip()
            nc = re.sub(r'\D', '', cv)
            if nc and len(nc) >= 4:
                existing_code_rows.append(r)

        template_row = hrow + 1
        styles = {}
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=template_row, column=col)
            styles[col] = {
                "font": copy(cell.font) if cell.font else None,
                "fill": copy(cell.fill) if cell.fill else None,
                "border": copy(cell.border) if cell.border else None,
                "alignment": copy(cell.alignment) if cell.alignment else None,
                "number_format": cell.number_format
            }

        if not existing_code_rows:
            count = len(user_codes)
            ws.cell(row=template_row, column=1, value=_get_code_str(user_codes[0]))
            if count > 1:
                ws.insert_rows(template_row + 1, count - 1)
                for j in range(i + 1, len(formato_rows)):
                    formato_rows[j] = (formato_rows[j][0] + count - 1, formato_rows[j][1])
            for idx in range(1, count):
                r = template_row + idx
                ws.cell(row=r, column=1, value=_get_code_str(user_codes[idx]))
                _apply_styles(ws, r, styles)
        else:
            for idx, code_row in enumerate(existing_code_rows):
                if idx < len(user_codes):
                    ws.cell(row=code_row, column=1, value=_get_code_str(user_codes[idx]))
                else:
                    ws.cell(row=code_row, column=1, value="")

            extra = len(user_codes) - len(existing_code_rows)
            if extra > 0:
                insert_at = existing_code_rows[-1] + 1
                ws.insert_rows(insert_at, extra)
                for idx in range(extra):
                    r = insert_at + idx
                    ws.cell(row=r, column=1, value=_get_code_str(user_codes[len(existing_code_rows) + idx]))
                    _apply_styles(ws, r, styles)


def _apply_styles(ws, row, styles):
    """Aplica estilos copiados a uma linha."""
    for col, style in styles.items():
        dst = ws.cell(row=row, column=col)
        if style.get("font"): dst.font = style["font"]
        if style.get("fill"): dst.fill = style["fill"]
        if style.get("border"): dst.border = style["border"]
        if style.get("alignment"): dst.alignment = style["alignment"]
        if style.get("number_format"): dst.number_format = style["number_format"]


def _find_code_data(codes_list, raw_val):
    """Procura raw_val em codes_list (strings ou objetos) e retorna (price, extra) individuais ou (None, None)."""
    for item in codes_list:
        code = item["code"] if isinstance(item, dict) else item
        if code.upper() == raw_val:
            if isinstance(item, dict):
                return item.get("price"), item.get("extra")
            break
    return None, None


def apply_prices(ws, mapping):
    """Percorre a planilha aplicando preços base + acréscimo nas colunas 2 e 3.
    
    Suporta preços individuais por código: se o código tiver price/extra próprio,
    usa esse valor. Caso contrário, usa o preço global do formato.
    """
    current_format = None
    current_price = None
    current_extra = None
    current_codes = []
    success_count = 0
    total_codes = 0

    for r in range(1, ws.max_row + 1):
        cell_val = str(ws.cell(row=r, column=1).value or "").strip().upper()

        dim = extract_dim_from_header(cell_val)
        if dim is not None:
            current_format = cell_val
            fmt_data = mapping.get(current_format, {})
            current_codes = fmt_data.get("codes", [])
            p = fmt_data.get("price")
            e = fmt_data.get("extra")
            if p is not None and e is not None and str(p) != "" and str(e) != "":
                try:
                    current_price = float(p)
                    current_extra = float(e)
                except (ValueError, TypeError):
                    current_price = None
                    current_extra = None
            else:
                current_price = None
                current_extra = None
            continue

        num_code = re.sub(r'\D', '', cell_val)
        if not num_code or len(num_code) < 4:
            continue

        total_codes += 1

        # Verifica se tem preço individual
        ind_price, ind_extra = _find_code_data(current_codes, cell_val)
        if ind_price is not None and ind_extra is not None:
            final_price = float(ind_price)
            final_extra = float(ind_extra)
        elif current_price is not None and current_extra is not None:
            final_price = current_price
            final_extra = current_extra
        else:
            continue

        set_price_cell(ws, r, 2, final_price)
        set_price_cell(ws, r, 3, final_price + final_extra)
        success_count += 1

    return success_count, total_codes


def merge_ai_products(formats_with_codes, ai_products):
    """Mescla produtos extraídos pela IA nos formatos existentes."""
    for prod in ai_products:
        tamanho = str(prod.get("tamanho", ""))
        codigo = str(prod.get("codigo", "")).strip()
        if not codigo:
            continue

        dim = extract_dim_from_header(tamanho)
        if dim is None:
            continue

        matched_format = None
        for fmt_header in formats_with_codes:
            fmt_dim = extract_dim_from_header(fmt_header)
            if fmt_dim == dim:
                matched_format = fmt_header
                break

        if matched_format and codigo not in formats_with_codes[matched_format]:
            formats_with_codes[matched_format].append(codigo)


def build_formats_list(formats_with_codes):
    """Converte o dict de formatos para lista ordenada para resposta JSON."""
    formats_list = []
    for fmt_header in sorted(formats_with_codes.keys()):
        formats_list.append({
            "header": fmt_header,
            "dim": extract_dim_from_header(fmt_header),
            "codes": formats_with_codes[fmt_header]
        })
    return formats_list


def save_excel_to_base64(wb):
    """Salva workbook em BytesIO e retorna base64."""
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return base64.b64encode(out.read()).decode('utf-8')